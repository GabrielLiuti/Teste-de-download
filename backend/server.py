from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import xmltodict
import io
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'fiscalmanager_secret_key_2025')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION = 24  # hours

security = HTTPBearer()

# Create the main app
app = FastAPI(title="FiscalManager Total API")
api_router = APIRouter(prefix="/api")

# ============= MODELS =============

class Usuario(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    email: EmailStr
    role: str = "usuario"  # admin, usuario
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UsuarioCreate(BaseModel):
    nome: str
    email: EmailStr
    senha: str
    role: Optional[str] = "usuario"

class UsuarioLogin(BaseModel):
    email: EmailStr
    senha: str

class Empresa(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nome: str
    cnpj: str
    rua: str
    numero: str
    bairro: str
    cidade: str
    estado: str
    cep: str
    regime_tributario: str  # Simples Nacional, Lucro Presumido, Lucro Real
    usuario_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmpresaCreate(BaseModel):
    nome: str
    cnpj: str
    rua: str
    numero: str
    bairro: str
    cidade: str
    estado: str
    cep: str
    regime_tributario: str

class Produto(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    nome: str
    codigo: str
    categoria: str
    valor_unitario: float
    aliquota_icms: float = 18.0
    aliquota_pis: float = 1.65
    aliquota_cofins: float = 7.6
    aliquota_ipi: float = 0.0
    usuario_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProdutoCreate(BaseModel):
    empresa_id: str
    nome: str
    codigo: str
    categoria: str
    valor_unitario: float
    aliquota_icms: Optional[float] = 18.0
    aliquota_pis: Optional[float] = 1.65
    aliquota_cofins: Optional[float] = 7.6
    aliquota_ipi: Optional[float] = 0.0

class ItemNF(BaseModel):
    produto_id: str
    produto_nome: str
    quantidade: float
    valor_unitario: float
    total_item: float
    icms: float
    pis: float
    cofins: float
    ipi: float

class NotaFiscal(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    empresa_id: str
    empresa_nome: str
    numero_nf: str
    data_emissao: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    itens: List[ItemNF]
    total_valor: float
    total_icms: float
    total_pis: float
    total_cofins: float
    total_ipi: float
    usuario_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class NotaFiscalCreate(BaseModel):
    empresa_id: str
    numero_nf: str
    itens: List[dict]

class DashboardStats(BaseModel):
    total_empresas: int
    total_produtos: int
    total_notas: int
    total_valor_notas: float
    total_impostos: dict
    notas_recentes: List[dict]

# ============= AUTHENTICATION =============

def create_token(usuario_id: str, email: str, role: str) -> str:
    payload = {
        'usuario_id': usuario_id,
        'email': email,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def verify_token(token: str) -> dict:
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    payload = verify_token(token)
    return payload

# ============= ROUTES - AUTH =============

@api_router.post("/auth/register", response_model=dict)
async def register(user: UsuarioCreate):
    # Check if email exists
    existing = await db.usuarios.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    # Hash password
    senha_hash = bcrypt.hashpw(user.senha.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    # Create user
    usuario = Usuario(
        nome=user.nome,
        email=user.email,
        role=user.role
    )
    
    doc = usuario.model_dump()
    doc['senha_hash'] = senha_hash
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.usuarios.insert_one(doc)
    
    token = create_token(usuario.id, usuario.email, usuario.role)
    
    return {
        "message": "Usuário cadastrado com sucesso",
        "token": token,
        "usuario": {
            "id": usuario.id,
            "nome": usuario.nome,
            "email": usuario.email,
            "role": usuario.role
        }
    }

@api_router.post("/auth/login", response_model=dict)
async def login(credentials: UsuarioLogin):
    # Find user
    user = await db.usuarios.find_one({"email": credentials.email})
    if not user:
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    
    # Verify password
    if not bcrypt.checkpw(credentials.senha.encode('utf-8'), user['senha_hash'].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    
    token = create_token(user['id'], user['email'], user['role'])
    
    return {
        "message": "Login realizado com sucesso",
        "token": token,
        "usuario": {
            "id": user['id'],
            "nome": user['nome'],
            "email": user['email'],
            "role": user['role']
        }
    }

# ============= ROUTES - EMPRESAS =============

@api_router.post("/empresas", response_model=Empresa)
async def create_empresa(empresa: EmpresaCreate, current_user: dict = Depends(get_current_user)):
    # Check if CNPJ already exists for this user
    existing = await db.empresas.find_one({"cnpj": empresa.cnpj, "usuario_id": current_user['usuario_id']})
    if existing:
        raise HTTPException(status_code=400, detail="CNPJ já cadastrado")
    
    empresa_obj = Empresa(**empresa.model_dump(), usuario_id=current_user['usuario_id'])
    doc = empresa_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.empresas.insert_one(doc)
    return empresa_obj

@api_router.get("/empresas", response_model=List[Empresa])
async def list_empresas(current_user: dict = Depends(get_current_user)):
    empresas = await db.empresas.find({"usuario_id": current_user['usuario_id']}, {"_id": 0}).to_list(1000)
    for e in empresas:
        if isinstance(e['created_at'], str):
            e['created_at'] = datetime.fromisoformat(e['created_at'])
    return empresas

@api_router.get("/empresas/{empresa_id}", response_model=Empresa)
async def get_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    empresa = await db.empresas.find_one({"id": empresa_id, "usuario_id": current_user['usuario_id']}, {"_id": 0})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    if isinstance(empresa['created_at'], str):
        empresa['created_at'] = datetime.fromisoformat(empresa['created_at'])
    return empresa

@api_router.put("/empresas/{empresa_id}", response_model=Empresa)
async def update_empresa(empresa_id: str, empresa_data: EmpresaCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.empresas.find_one({"id": empresa_id, "usuario_id": current_user['usuario_id']})
    if not existing:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    update_data = empresa_data.model_dump()
    await db.empresas.update_one({"id": empresa_id}, {"$set": update_data})
    
    updated = await db.empresas.find_one({"id": empresa_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/empresas/{empresa_id}")
async def delete_empresa(empresa_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.empresas.delete_one({"id": empresa_id, "usuario_id": current_user['usuario_id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return {"message": "Empresa excluída com sucesso"}

# ============= ROUTES - PRODUTOS =============

@api_router.post("/produtos", response_model=Produto)
async def create_produto(produto: ProdutoCreate, current_user: dict = Depends(get_current_user)):
    # Verify empresa belongs to user
    empresa = await db.empresas.find_one({"id": produto.empresa_id, "usuario_id": current_user['usuario_id']})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    produto_obj = Produto(**produto.model_dump(), usuario_id=current_user['usuario_id'])
    doc = produto_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.produtos.insert_one(doc)
    return produto_obj

@api_router.get("/produtos", response_model=List[Produto])
async def list_produtos(empresa_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"usuario_id": current_user['usuario_id']}
    if empresa_id:
        query["empresa_id"] = empresa_id
    
    produtos = await db.produtos.find(query, {"_id": 0}).to_list(1000)
    for p in produtos:
        if isinstance(p['created_at'], str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return produtos

@api_router.get("/produtos/{produto_id}", response_model=Produto)
async def get_produto(produto_id: str, current_user: dict = Depends(get_current_user)):
    produto = await db.produtos.find_one({"id": produto_id, "usuario_id": current_user['usuario_id']}, {"_id": 0})
    if not produto:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    if isinstance(produto['created_at'], str):
        produto['created_at'] = datetime.fromisoformat(produto['created_at'])
    return produto

@api_router.put("/produtos/{produto_id}", response_model=Produto)
async def update_produto(produto_id: str, produto_data: ProdutoCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.produtos.find_one({"id": produto_id, "usuario_id": current_user['usuario_id']})
    if not existing:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    
    update_data = produto_data.model_dump()
    await db.produtos.update_one({"id": produto_id}, {"$set": update_data})
    
    updated = await db.produtos.find_one({"id": produto_id}, {"_id": 0})
    if isinstance(updated['created_at'], str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return updated

@api_router.delete("/produtos/{produto_id}")
async def delete_produto(produto_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.produtos.delete_one({"id": produto_id, "usuario_id": current_user['usuario_id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return {"message": "Produto excluído com sucesso"}

# ============= ROUTES - NOTAS FISCAIS =============

def calcular_impostos(item: dict, regime: str) -> dict:
    """Calcula impostos de um item"""
    valor_total_item = item['valor_unitario'] * item['quantidade']
    
    icms = valor_total_item * (item.get('aliquota_icms', 18.0) / 100)
    pis = valor_total_item * (item.get('aliquota_pis', 1.65) / 100)
    cofins = valor_total_item * (item.get('aliquota_cofins', 7.6) / 100)
    ipi = valor_total_item * (item.get('aliquota_ipi', 0.0) / 100)
    
    # Simples Nacional tem redução de ICMS
    if regime == 'Simples Nacional':
        icms *= 0.7  # 30% de redução
    
    return {
        'icms': round(icms, 2),
        'pis': round(pis, 2),
        'cofins': round(cofins, 2),
        'ipi': round(ipi, 2)
    }

@api_router.post("/notas", response_model=NotaFiscal)
async def create_nota(nota: NotaFiscalCreate, current_user: dict = Depends(get_current_user)):
    # Verify empresa
    empresa = await db.empresas.find_one({"id": nota.empresa_id, "usuario_id": current_user['usuario_id']})
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    
    # Calculate taxes for each item
    itens_calculados = []
    total_valor = 0
    total_icms = 0
    total_pis = 0
    total_cofins = 0
    total_ipi = 0
    
    for item_data in nota.itens:
        # Get produto details
        produto = await db.produtos.find_one({"id": item_data['produto_id'], "usuario_id": current_user['usuario_id']})
        if not produto:
            raise HTTPException(status_code=404, detail=f"Produto {item_data['produto_id']} não encontrado")
        
        # Merge produto data with item data
        item_completo = {
            'produto_id': produto['id'],
            'produto_nome': produto['nome'],
            'quantidade': item_data['quantidade'],
            'valor_unitario': produto['valor_unitario'],
            'aliquota_icms': produto['aliquota_icms'],
            'aliquota_pis': produto['aliquota_pis'],
            'aliquota_cofins': produto['aliquota_cofins'],
            'aliquota_ipi': produto['aliquota_ipi']
        }
        
        # Calculate
        impostos = calcular_impostos(item_completo, empresa['regime_tributario'])
        total_item = item_completo['valor_unitario'] * item_completo['quantidade']
        
        item_nf = ItemNF(
            produto_id=item_completo['produto_id'],
            produto_nome=item_completo['produto_nome'],
            quantidade=item_completo['quantidade'],
            valor_unitario=item_completo['valor_unitario'],
            total_item=round(total_item, 2),
            icms=impostos['icms'],
            pis=impostos['pis'],
            cofins=impostos['cofins'],
            ipi=impostos['ipi']
        )
        
        itens_calculados.append(item_nf)
        total_valor += item_nf.total_item
        total_icms += item_nf.icms
        total_pis += item_nf.pis
        total_cofins += item_nf.cofins
        total_ipi += item_nf.ipi
    
    # Create nota fiscal
    nota_fiscal = NotaFiscal(
        empresa_id=nota.empresa_id,
        empresa_nome=empresa['nome'],
        numero_nf=nota.numero_nf,
        itens=itens_calculados,
        total_valor=round(total_valor, 2),
        total_icms=round(total_icms, 2),
        total_pis=round(total_pis, 2),
        total_cofins=round(total_cofins, 2),
        total_ipi=round(total_ipi, 2),
        usuario_id=current_user['usuario_id']
    )
    
    doc = nota_fiscal.model_dump()
    doc['data_emissao'] = doc['data_emissao'].isoformat()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['itens'] = [item.model_dump() for item in nota_fiscal.itens]
    
    await db.notas_fiscais.insert_one(doc)
    return nota_fiscal

@api_router.get("/notas", response_model=List[NotaFiscal])
async def list_notas(empresa_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"usuario_id": current_user['usuario_id']}
    if empresa_id:
        query["empresa_id"] = empresa_id
    
    notas = await db.notas_fiscais.find(query, {"_id": 0}).to_list(1000)
    for n in notas:
        if isinstance(n['data_emissao'], str):
            n['data_emissao'] = datetime.fromisoformat(n['data_emissao'])
        if isinstance(n['created_at'], str):
            n['created_at'] = datetime.fromisoformat(n['created_at'])
    return notas

@api_router.get("/notas/{nota_id}", response_model=NotaFiscal)
async def get_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    nota = await db.notas_fiscais.find_one({"id": nota_id, "usuario_id": current_user['usuario_id']}, {"_id": 0})
    if not nota:
        raise HTTPException(status_code=404, detail="Nota fiscal não encontrada")
    if isinstance(nota['data_emissao'], str):
        nota['data_emissao'] = datetime.fromisoformat(nota['data_emissao'])
    if isinstance(nota['created_at'], str):
        nota['created_at'] = datetime.fromisoformat(nota['created_at'])
    return nota

@api_router.delete("/notas/{nota_id}")
async def delete_nota(nota_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.notas_fiscais.delete_one({"id": nota_id, "usuario_id": current_user['usuario_id']})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Nota fiscal não encontrada")
    return {"message": "Nota fiscal excluída com sucesso"}

# ============= ROUTES - DASHBOARD =============

@api_router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard(current_user: dict = Depends(get_current_user)):
    # Count totals
    total_empresas = await db.empresas.count_documents({"usuario_id": current_user['usuario_id']})
    total_produtos = await db.produtos.count_documents({"usuario_id": current_user['usuario_id']})
    total_notas = await db.notas_fiscais.count_documents({"usuario_id": current_user['usuario_id']})
    
    # Sum totals from notas
    notas = await db.notas_fiscais.find({"usuario_id": current_user['usuario_id']}, {"_id": 0}).to_list(1000)
    
    total_valor_notas = sum(n.get('total_valor', 0) for n in notas)
    total_impostos = {
        'icms': sum(n.get('total_icms', 0) for n in notas),
        'pis': sum(n.get('total_pis', 0) for n in notas),
        'cofins': sum(n.get('total_cofins', 0) for n in notas),
        'ipi': sum(n.get('total_ipi', 0) for n in notas)
    }
    
    # Get recent notas
    notas_recentes = sorted(notas, key=lambda x: x.get('data_emissao', ''), reverse=True)[:5]
    
    return DashboardStats(
        total_empresas=total_empresas,
        total_produtos=total_produtos,
        total_notas=total_notas,
        total_valor_notas=round(total_valor_notas, 2),
        total_impostos=total_impostos,
        notas_recentes=notas_recentes
    )

# ============= ROUTES - RELATÓRIOS =============

@api_router.get("/relatorios/pdf")
async def gerar_relatorio_pdf(current_user: dict = Depends(get_current_user)):
    notas = await db.notas_fiscais.find({"usuario_id": current_user['usuario_id']}, {"_id": 0}).to_list(1000)
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title = Paragraph("<b>Relatório Fiscal - FiscalManager Total</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # Summary
    total_valor = sum(n.get('total_valor', 0) for n in notas)
    total_icms = sum(n.get('total_icms', 0) for n in notas)
    total_pis = sum(n.get('total_pis', 0) for n in notas)
    total_cofins = sum(n.get('total_cofins', 0) for n in notas)
    total_ipi = sum(n.get('total_ipi', 0) for n in notas)
    
    summary_data = [
        ['Total de Notas', str(len(notas))],
        ['Valor Total', f'R$ {total_valor:,.2f}'],
        ['ICMS', f'R$ {total_icms:,.2f}'],
        ['PIS', f'R$ {total_pis:,.2f}'],
        ['COFINS', f'R$ {total_cofins:,.2f}'],
        ['IPI', f'R$ {total_ipi:,.2f}']
    ]
    
    summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(buffer, media_type="application/pdf", headers={
        "Content-Disposition": "attachment; filename=relatorio_fiscal.pdf"
    })

@api_router.get("/relatorios/excel")
async def gerar_relatorio_excel(current_user: dict = Depends(get_current_user)):
    notas = await db.notas_fiscais.find({"usuario_id": current_user['usuario_id']}, {"_id": 0}).to_list(1000)
    
    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Fiscal"
    
    # Headers
    headers = ['Número NF', 'Empresa', 'Data Emissão', 'Valor Total', 'ICMS', 'PIS', 'COFINS', 'IPI']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for nota in notas:
        ws.append([
            nota.get('numero_nf', ''),
            nota.get('empresa_nome', ''),
            nota.get('data_emissao', ''),
            nota.get('total_valor', 0),
            nota.get('total_icms', 0),
            nota.get('total_pis', 0),
            nota.get('total_cofins', 0),
            nota.get('total_ipi', 0)
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        "Content-Disposition": "attachment; filename=relatorio_fiscal.xlsx"
    })

# ============= MIDDLEWARE & STARTUP =============

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

@api_router.get("/")
async def root():
    return {"message": "FiscalManager Total API - v1.0"}